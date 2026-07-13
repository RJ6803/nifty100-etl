from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ScreenerExporter:

    HEADER_FILL = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
    )

    PASS_FILL = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid",
    )

    FAIL_FILL = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid",
    )

    DARK_GREEN = PatternFill(
        start_color="006100",
        end_color="006100",
        fill_type="solid",
    )

    GREEN = PatternFill(
        start_color="00B050",
        end_color="00B050",
        fill_type="solid",
    )

    LIGHT_GREEN = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid",
    )

    YELLOW = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid",
    )

    RED = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid",
    )

    def export(self, results, filename, config):

        wb = Workbook()

        wb.remove(wb.active)

        for preset_name, df in results.items():

            columns = [
                "company_id",
                "company_name",
                "broad_sector",
                "return_on_equity_pct",
                "return_on_capital_employed_pct",
                "net_profit_margin_pct",
                "debt_to_equity",
                "interest_coverage",
                "free_cash_flow_cr",
                "revenue_cagr_5yr",
                "pat_cagr_5yr",
                "eps_cagr_5yr",
                "asset_turnover",
                "market_cap_crore",
                "sales",
                "net_profit",
                "pe_ratio",
                "pb_ratio",
                "dividend_yield_pct",
                "dividend_payout_ratio_pct",
                "composite_quality_score",
                "quality_rank",
            ]

            existing = [c for c in columns if c in df.columns]

            df = df[existing]

            ws = wb.create_sheet(title=preset_name[:31])

            count = len(df)

            avg_score = (
                round(df["composite_quality_score"].mean(), 2)
                if "composite_quality_score" in df.columns
                else ""
            )

            top_company = (
                df.iloc[0]["company_name"]
                if count > 0
                else ""
            )

            # ---------------- HEADER ----------------

            for col, column in enumerate(df.columns, start=1):

                cell = ws.cell(row=1, column=col)

                cell.value = column

                cell.fill = self.HEADER_FILL

                cell.font = self.HEADER_FONT

            rules = config.get(preset_name, {})

            # ---------------- DATA ----------------

            for r, row in enumerate(df.itertuples(index=False), start=2):

                for c, value in enumerate(row, start=1):

                    cell = ws.cell(row=r, column=c)

                    cell.value = value

                    column_name = df.columns[c - 1]

                    fill = None

                    if value is None:
                        continue

                    # Composite score coloring

                    if column_name == "composite_quality_score":

                        if value >= 90:
                            fill = self.DARK_GREEN

                        elif value >= 75:
                            fill = self.GREEN

                        elif value >= 60:
                            fill = self.LIGHT_GREEN

                        elif value >= 40:
                            fill = self.YELLOW

                        else:
                            fill = self.RED

                    # Rule highlighting

                    elif column_name == "return_on_equity_pct" and "roe_min" in rules:
                        fill = self.PASS_FILL if value >= rules["roe_min"] else self.FAIL_FILL

                    elif column_name == "sales" and "sales_min" in rules:
                        fill = self.PASS_FILL if value >= rules["sales_min"] else self.FAIL_FILL

                    elif column_name == "net_profit" and "net_profit_min" in rules:
                        fill = self.PASS_FILL if value >= rules["net_profit_min"] else self.FAIL_FILL

                    elif column_name == "asset_turnover" and "asset_turnover_min" in rules:
                        fill = self.PASS_FILL if value >= rules["asset_turnover_min"] else self.FAIL_FILL

                    elif column_name == "eps_cagr_5yr" and "eps_cagr_5yr_min" in rules:
                        fill = self.PASS_FILL if value >= rules["eps_cagr_5yr_min"] else self.FAIL_FILL

                    elif (
                        column_name == "operating_profit_margin_pct"
                        and "operating_profit_margin_min" in rules
                    ):
                        fill = (
                            self.PASS_FILL
                            if value >= rules["operating_profit_margin_min"]
                            else self.FAIL_FILL
                        )

                    elif (
                        column_name == "dividend_payout_ratio_pct"
                        and "dividend_payout_max" in rules
                    ):
                        fill = (
                            self.PASS_FILL
                            if value <= rules["dividend_payout_max"]
                            else self.FAIL_FILL
                        )

                    elif column_name == "debt_to_equity" and "debt_to_equity_max" in rules:
                        fill = (
                            self.PASS_FILL
                            if value <= rules["debt_to_equity_max"]
                            else self.FAIL_FILL
                        )

                    elif column_name == "free_cash_flow_cr" and "free_cash_flow_min" in rules:
                        fill = (
                            self.PASS_FILL
                            if value >= rules["free_cash_flow_min"]
                            else self.FAIL_FILL
                        )

                    elif (
                        column_name == "revenue_cagr_5yr"
                        and "revenue_cagr_5yr_min" in rules
                    ):
                        fill = (
                            self.PASS_FILL
                            if value >= rules["revenue_cagr_5yr_min"]
                            else self.FAIL_FILL
                        )

                    elif column_name == "pat_cagr_5yr" and "pat_cagr_5yr_min" in rules:
                        fill = (
                            self.PASS_FILL
                            if value >= rules["pat_cagr_5yr_min"]
                            else self.FAIL_FILL
                        )

                    elif column_name == "pe_ratio" and "pe_max" in rules:
                        fill = (
                            self.PASS_FILL
                            if value <= rules["pe_max"]
                            else self.FAIL_FILL
                        )

                    elif column_name == "pb_ratio" and "pb_max" in rules:
                        fill = (
                            self.PASS_FILL
                            if value <= rules["pb_max"]
                            else self.FAIL_FILL
                        )

                    elif (
                        column_name == "dividend_yield_pct"
                        and "dividend_yield_min" in rules
                    ):
                        fill = (
                            self.PASS_FILL
                            if value >= rules["dividend_yield_min"]
                            else self.FAIL_FILL
                        )

                    elif (
                        column_name == "market_cap_crore"
                        and "market_cap_min" in rules
                    ):
                        fill = (
                            self.PASS_FILL
                            if value >= rules["market_cap_min"]
                            else self.FAIL_FILL
                        )

                    elif (
                        column_name == "interest_coverage"
                        and "interest_coverage_min" in rules
                    ):
                        fill = (
                            self.PASS_FILL
                            if value >= rules["interest_coverage_min"]
                            else self.FAIL_FILL
                        )

                    if fill:
                        cell.fill = fill

            # ---------------- AUTO WIDTH ----------------

            for column_cells in ws.columns:

                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )

                ws.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = min(length + 2, 35)

            ws.freeze_panes = "A2"

        wb.save(filename)

        print(f"\nWorkbook saved -> {filename}")
