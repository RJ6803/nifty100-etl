from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelFormatter:

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)

    def format_all_sheets(self):

        for ws in self.workbook.worksheets:

            # Freeze header
            ws.freeze_panes = "A2"

            # Enable filter
            ws.auto_filter.ref = ws.dimensions

            # Header formatting
            header_fill = PatternFill(
                fill_type="solid",
                start_color="1F4E78",
                end_color="1F4E78"
            )

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # Auto column width
            for column in ws.columns:

                length = 0

                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        length = max(length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[column_letter].width = min(length + 3, 45)

        self.workbook.save(self.file_path)

        print("Excel formatting completed.")