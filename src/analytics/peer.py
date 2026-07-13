"""Peer-group percentile calculations, radar charts and Excel reporting."""

from __future__ import annotations

import sqlite3
from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # Reports are generated in headless ETL environments.
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.screener.engine import ScreenerEngine


RANK_METRICS = {
    "ROE": "return_on_equity_pct", "ROCE": "return_on_capital_employed_pct",
    "Net Profit Margin": "net_profit_margin_pct", "Debt to Equity": "debt_to_equity",
    "Free Cash Flow": "free_cash_flow_cr", "PAT CAGR 5yr": "pat_cagr_5yr",
    "Revenue CAGR 5yr": "revenue_cagr_5yr", "EPS CAGR 5yr": "eps_cagr_5yr",
    "Interest Coverage": "interest_coverage", "Asset Turnover": "asset_turnover",
}
REPORT_COLUMNS = [
    "company_id", "company_name", "broad_sector", "return_on_equity_pct",
    "return_on_capital_employed_pct", "net_profit_margin_pct", "operating_profit_margin_pct",
    "debt_to_equity", "interest_coverage", "free_cash_flow_cr", "cash_from_operations_cr",
    "cfo_pat_ratio", "revenue_cagr_3yr", "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "asset_turnover", "market_cap_crore", "sales", "net_profit", "pe_ratio", "pb_ratio",
    "dividend_yield_pct", "dividend_payout_ratio_pct", "composite_quality_score",
]


class PeerComparisonEngine:
    def __init__(self, db_path="data/nifty100.db", peer_groups_path="data/raw/peer_groups.xlsx"):
        self.db_path = Path(db_path)
        self.groups = pd.read_excel(peer_groups_path)
        required = {"peer_group_name", "company_id", "is_benchmark"}
        missing = required - set(self.groups.columns)
        if missing:
            raise ValueError(f"peer_groups.xlsx is missing columns: {sorted(missing)}")
        self.data = ScreenerEngine(db_path=self.db_path).df
        self.percentiles = pd.DataFrame()

    @staticmethod
    def _percent_rank(series, inverse=False):
        values = pd.to_numeric(series.replace("Debt Free", np.inf), errors="coerce")
        valid = values.notna()
        n = valid.sum()
        result = pd.Series(np.nan, index=series.index, dtype=float)
        if n == 1:
            result.loc[valid] = 1.0
        elif n > 1:
            result.loc[valid] = (values.loc[valid].rank(method="min") - 1) / (n - 1)
        return 1 - result if inverse else result

    def compute_percentiles(self):
        joined = self.groups.merge(self.data, on="company_id", how="left", validate="many_to_one")
        records = []
        for group, frame in joined.groupby("peer_group_name", sort=True):
            # Some peer-company identifiers are absent from a historical ratio
            # extract; keep the rank record traceable instead of dropping it.
            year = str(frame["year"].dropna().iloc[0]) if frame["year"].notna().any() else "latest"
            for metric, column in RANK_METRICS.items():
                ranks = self._percent_rank(frame[column], inverse=(column == "debt_to_equity"))
                for index, row in frame.iterrows():
                    records.append({"company_id": row.company_id, "peer_group_name": group, "metric": metric,
                                    "value": pd.to_numeric(pd.Series([row[column]]), errors="coerce").iloc[0],
                                    "percentile_rank": ranks.loc[index], "year": year})
        self.percentiles = pd.DataFrame(records)
        return self.percentiles

    def persist_percentiles(self):
        if self.percentiles.empty:
            self.compute_percentiles()
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""CREATE TABLE IF NOT EXISTS peer_percentiles (
                company_id TEXT NOT NULL, peer_group_name TEXT NOT NULL, metric TEXT NOT NULL,
                value REAL, percentile_rank REAL, year TEXT NOT NULL,
                PRIMARY KEY (company_id, peer_group_name, metric, year))""")
            conn.execute("DELETE FROM peer_percentiles")
            self.percentiles.to_sql("peer_percentiles", conn, if_exists="append", index=False)

    def peer_group_for(self, company_id):
        assigned = self.groups.loc[self.groups.company_id.eq(company_id), "peer_group_name"].tolist()
        return assigned[0] if assigned else "No peer group assigned"

    def _group_frame(self, group):
        frame = self.groups.loc[self.groups.peer_group_name.eq(group)].merge(self.data, on="company_id", how="left")
        ranks = self.percentiles.loc[self.percentiles.peer_group_name.eq(group)].pivot(
            index="company_id", columns="metric", values="percentile_rank"
        ).rename(columns=lambda x: f"{x} percentile")
        return frame.merge(ranks, left_on="company_id", right_index=True, how="left")

    def export_excel(self, output_path="output/peer_comparison.xlsx"):
        if self.percentiles.empty:
            self.compute_percentiles()
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook(); wb.remove(wb.active)
        header = PatternFill("solid", fgColor="1F4E78")
        amber = PatternFill("solid", fgColor="FFD966")
        green = PatternFill("solid", fgColor="C6EFCE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        red = PatternFill("solid", fgColor="FFC7CE")
        for group in sorted(self.groups.peer_group_name.unique()):
            frame = self._group_frame(group)
            cols = [c for c in REPORT_COLUMNS if c in frame.columns] + [f"{m} percentile" for m in RANK_METRICS]
            ws = wb.create_sheet(group[:31]); ws.append(cols)
            for cell in ws[1]: cell.fill, cell.font = header, Font(bold=True, color="FFFFFF")
            for _, row in frame.iterrows():
                ws.append([row.get(c) for c in cols])
                if bool(row.get("is_benchmark", False)):
                    for cell in ws[ws.max_row]: cell.fill = amber
            # Per-metric median, never a percentile median masquerading as a company.
            ws.append(["", "Peer group median"] + [frame[c].median() if c in frame and pd.api.types.is_numeric_dtype(frame[c]) else None for c in cols[2:]])
            for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            for col in range(1, ws.max_column + 1):
                letter = get_column_letter(col); ws.column_dimensions[letter].width = min(28, max(12, max(len(str(c.value or "")) for c in ws[letter]) + 2))
            for col in range(1, ws.max_column + 1):
                if "percentile" in str(ws.cell(1, col).value):
                    ws.conditional_formatting.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row - 1}", CellIsRule(operator="greaterThanOrEqual", formula=["0.75"], fill=green))
                    ws.conditional_formatting.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row - 1}", CellIsRule(operator="between", formula=["0.25", "0.749999"], fill=yellow))
                    ws.conditional_formatting.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row - 1}", CellIsRule(operator="lessThanOrEqual", formula=["0.25"], fill=red))
            ws.freeze_panes = "A2"
        wb.save(output_path)

    def generate_radar_charts(self, directory="reports/radar_charts"):
        if self.percentiles.empty:
            self.compute_percentiles()
        directory = Path(directory); directory.mkdir(parents=True, exist_ok=True)
        radar_metrics = ["ROE", "ROCE", "Net Profit Margin", "Debt to Equity", "Free Cash Flow", "PAT CAGR 5yr", "Revenue CAGR 5yr"]
        global_scores = self.data.set_index("company_id")["composite_quality_score"]
        for company_id in self.data.company_id.dropna().unique():
            group = self.peer_group_for(company_id)
            if group == "No peer group assigned":
                frame, title = self.data, f"{company_id} vs Nifty 100 average"
            else:
                frame, title = self._group_frame(group), f"{company_id} vs {group} average"
            company = frame.loc[frame.company_id.eq(company_id)].iloc[0]
            values, average, labels = [], [], radar_metrics + ["Composite Score"]
            for metric in radar_metrics:
                ranks = self.percentiles.loc[(self.percentiles.peer_group_name.eq(group)) & (self.percentiles.metric.eq(metric)), ["company_id", "percentile_rank"]]
                if group == "No peer group assigned":
                    column = RANK_METRICS[metric]; value = pd.to_numeric(pd.Series([company[column]]), errors="coerce").iloc[0]
                    lo, hi = pd.to_numeric(self.data[column], errors="coerce").quantile([.1, .9]); score = 50 if hi == lo else np.clip((value-lo)*100/(hi-lo), 0, 100)
                    values.append(score); average.append(50)
                else:
                    values.append(float(ranks.loc[ranks.company_id.eq(company_id), "percentile_rank"].iloc[0] * 100)); average.append(ranks.percentile_rank.mean() * 100)
            values.append(float(global_scores.get(company_id, 0))); average.append(float(frame.composite_quality_score.mean()))
            angles = np.linspace(0, 2*np.pi, len(labels), endpoint=False).tolist(); values += values[:1]; average += average[:1]; angles += angles[:1]
            fig, ax = plt.subplots(figsize=(7, 7), subplot_kw={"polar": True})
            ax.plot(angles, values, color="#1976D2", linewidth=2, label=company_id); ax.fill(angles, values, color="#1976D2", alpha=.22)
            ax.plot(angles, average, color="#E67E22", linewidth=2, linestyle="--", label="Peer average")
            ax.set_xticks(angles[:-1]); ax.set_xticklabels(labels, fontsize=9); ax.set_ylim(0, 100); ax.set_title(title, pad=24, fontsize=12); ax.legend(loc="upper right", bbox_to_anchor=(1.25, 1.15))
            fig.tight_layout(); fig.savefig(directory / f"{company_id}_radar.png", dpi=150); plt.close(fig)

    def run(self):
        self.compute_percentiles(); self.persist_percentiles(); self.export_excel(); self.generate_radar_charts()
        return self.percentiles


if __name__ == "__main__":
    PeerComparisonEngine().run()
